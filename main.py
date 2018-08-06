from bs4 import BeautifulSoup
import requests
import re
import openpyxl


icp_test_pages = ["lowyat.net","renren.com","paultan.org","mudah.my","moe.gov.my","hide.me","uitm.edu.my","hasil.gov.my","hongleongconnect.my","poslaju.com.my","1govuc.gov.my","um.edu.my","cari.com.my","upm.edu.my","ukm.my","orientaldaily.com.my","kwsp.gov.my","thrivethemes.com","utm.my","rtm.gov.my","iium.edu.my","eghrmis.gov.my","spa.gov.my","moh.gov.my","kwongwah.com.my","uum.edu.my","unimas.my","ums.edu.my","hlb.com.my","tm.com.my","wanista.com","upsi.edu.my","maxis.com.my","utar.edu.my","ptptn.gov.my","uthm.edu.my","mohe.gov.my","hidden-street.net","maybank.com","ipay88.com.my","irakyat.com.my","umt.edu.my","ump.edu.my","cinema.com.my","uob.com.my","cimbbank.com.my","customs.gov.my","bursamalaysia.com","imi.gov.my","dbp.gov.my","technave.com","seagm.com","perkeso.gov.my","myeg.com.my","mobe.com","airasiago.com.my","cimb.com","keyauto.my","sarawak.gov.my","jpa.gov.my","ktmb.com.my","yes.my","hideystudio.com","dosh.gov.my","oum.edu.my","unimap.edu.my","busonlineticket.com","bnm.gov.my","malaysiastock.biz","touchngo.com.my","utem.edu.my","unikl.edu.my","newinti.edu.my","revenue.com.my","unisza.edu.my","mediu.edu.my","gdexpress.com","usim.edu.my","jpj.gov.my","xmu.edu.my","bankislam.com.my","easyparcel.my","bernama.com","imoney.my","onlinepayment.com.my","affinonline.com","themalaymailonline.com","sportstoto.com.my","mdec.com.my","petronas.com.my","educationmalaysia.gov.my","ssm.com.my","zapmeta.com.my","kln.gov.my","perodua.com.my","cidb.gov.my","ringgitplus.com","taylors.edu.my","sabah.gov.my","publicmutual.com.my","items-ipgm.edu.my","ambank.com.my","mara.gov.my","rhbtradesmart.com","redone.com.my","financial-link.com.my","hoto.com","utp.edu.my","selangor.gov.my","ambankgroup.com","terengganu.gov.my","magnum4d.my","innity.com","vanguardelectronic.com.my","treasury.gov.my","e-solat.gov.my","mpm.edu.my","jobsmalaysia.gov.my","freejapanesefont.com","blinkwebinars.com","itc.gov.my","hyipartner.com","islam.gov.my","malaysiafreebies.com","prudential.com.my","malaysia.gov.my","rmp.gov.my","penerangan.gov.my","homepro.com.my","wvcentral.org","politeknik.edu.my","e-ghl.com","mbocinemas.com","solidpartners.com.my","abxexpress.com.my","ppimalaysia.com","simlecco.com.my","pith.com.my","time.com.my","myhealth.gov.my","klia2.info","radiomalaysia.net","mpob.gov.my","alliancebank.com.my","umk.edu.my","bumiarmada.com","proaudiotorrents.org","toyota.com.my","cimb-bizchannel.com.my","azlyricdb.com","directd.com.my","myrapid.com.my","ispring.com.my","msu.edu.my","uniten.edu.my","rcdreamz.com.my","serveron.com.my","nottingham.edu.my","tattgiap.com.my","alleights.com.my","biolife.com.my","johor.gov.my","clefdisplay.com","affinbank.com.my","kuis.edu.my","agc.gov.my","malayalasangeetham.info","mychef.com.my","hostingchecker.com","adk.gov.my","mampu.gov.my","sportsaffairs.com.my","sarawakborneotour.com","yli.com.my","citylinkexpress.com","kpkt.gov.my","vbtutor.net","mir.com.my","kpypj.edu.my","skynet.com.my","jkm.gov.my","curtin.edu.my","bonuslink.com.my","perak.gov.my","jbtalks.cc","penang.gov.my","onexox.my","ucsiuniversity.edu.my","youbeli.com","ismaweb.net","jpapencen.gov.my","doappx.com","proton.com","jmtop.com","asnb.com.my","bomba.gov.my","malaysianbar.org.my","3ecpa.com.my","bitspyder.net","moha.gov.my","unitar.my","bloc-48.com","cidos.edu.my","studymalaysia.com","cikguhailmi.com","hlebroking.com","monash.edu.my","sarawakenergy.com.my","apiit.edu.my","ssm-einfo.my","mqa.gov.my","jpkk.edu.my","pix.my","mmog.asia","merimen.com","surah.my","excard.com.my","prubsn.com.my","hype.my","mykamus.com","jkr.gov.my","dgvcl.com","muamalat.com.my","crazyartzone.com","marykayintouch.com.my","doe.gov.my","matrik.edu.my","jupem.gov.my","tranungkite.net","jais.gov.my","mgvcl.com","kptm.edu.my","xox.com.my","dbkl.gov.my","iloginhr.com","eterminal.net.my","kwp.gov.my","mystarjob.com","poliku.edu.my","sprm.gov.my","webcash.com.my","fomema.com.my","shashinki.com","tbsbts.com.my","b2b.com.my","jompay.com.my","onlinetuition.com.my","extremebb.net","nbmcw.com","kliaekspres.com","mod.gov.my","calculator.com.my","mypostonline.com.my","oto.my","syabas.com.my","masteryacademy.com","e-print.my","citcat.com","publicgold.com.my","unisel.edu.my","greateasterntakaful.com","tourism.gov.my","chartnexus.com","swinburne.edu.my","polikk.edu.my","fragoimpex.com","onlinewe.net","nissan.com.my","yellavia.com","jpm.gov.my","jestineyong.com","pahang.gov.my","rhbinvest.com","hrdf.com.my","ftms.edu.my","cgso.gov.my","jpj.my","audit.gov.my","d-synergy.com","bless.gov.my","zakatselangor.com.my","addgadgets.com","plus.com.my","kbs.gov.my","sppim.gov.my","apu.edu.my","dineshbakshi.com","labuanibfc.com","mimos.my","bicyclebuysell.com","upnm.edu.my","agronet.com.my","arccjournals.com","careerjet.com.my","hsbcamanah.com.my","lphs.gov.my","panduanmalaysia.com","hla.com.my","kelantan.gov.my","hcsb.com.my","pas.org.my","jtm.gov.my","spab.gov.my","atomx.com","basaer-online.com","sunwaylagoon.com","bigsale.com.my","colourink.com.my","kkr.gov.my","giant.com.my","mmallv2u.com","matrade.gov.my","segi.edu.my","spad.gov.my","tpb.com.my","moa.gov.my","zikrihusaini.com","ikimfm.my","mardi.gov.my","mynic.my","ppa1m.gov.my","advenueplatform.com","iab.edu.my","iic.edu.my","hoopsstation.com","takaful-malaysia.com.my","aeonretail.com.my","toypanic.com","mii4u.org","umno-online.my","rspo.org","hostelhunting.com","epson.com.my","e-cover.com.my","ummc.edu.my","cibmall.net","canon.com.my","mediline.com.my","iukl.edu.my","azamtv.com","polimelaka.edu.my","klangwesley.com","ipserverone.info","redappletravel.com","vectorise.net","avls.com.my","lap.com.my","unifi.my","mbfcards.com","iwk.com.my","cuckoo.com.my","bigsweep.com.my","fxdailyinfo.com","rurallink.gov.my","zerotohundred.com","halal.gov.my","tiendeo.my","mafhq.mil.my","psas.edu.my","puo.edu.my","mcmc.gov.my","icu.gov.my","edmarker.com","investalks.com","msd.net.my","mface.me","polisas.edu.my","smecorp.gov.my","sirim.my","mix.fm","kpjhealth.com.my","javadhotel.com","checkpointspot.asia","andrewchoo.edu.my","data.gov.my","fmm.org.my","sesb.com.my","lgm.gov.my","dvs.gov.my","fwcms.com.my","cmctos.com.my","allithypermarket.com.my","ns.gov.my","impulse.com.my","chilis.com.my","graduan.com","fitproconnect.com","domikado.com","taonline.com.my","aeu.edu.my","dressingpaula.com","niosh.com.my","townplan.gov.my","universitymalaysia.net","bph.gov.my","popular.com.my","yp.org.my","northport.com.my","lipstiq.com","fama.gov.my","aeonbig.com.my","eauto.my","mot.gov.my","accprint4u.com","st.gov.my","volkswagen.com.my","central-hosting.com","ppdpgudang.edu.my","evergreengroup.com.my","rhbinsurance.com.my","mylabourlaw.net","jobstock.com.my","psen.ir","psp.edu.my","mdec.my","kldslr.com","mydin.com.my","kvc.com.my","y5zone.my","joeyyap.com","ace.net.my","trademysuperbike.com.my","litefm.com.my","hitb.org","nationwide2u.com","ecoworld.my","cgnarzuki.com","tateyama.com.my","cubizone.com","hlisb.com.my","arkib.gov.my","imaanboutique.com","zoom-a.com","simedarby.com","mybest.com.my","hkl.gov.my","twitraining.com","myipo.gov.my","businesslist.my","jpnns.gov.my","pharmacy.gov.my","worldfishcenter.org","nettium.net","akpk.org.my","kclau.com","mm2h.gov.my","themalaysiantimes.com.my","mia.org.my","dziennik-polityczny.com","benpartners.com","propertyking.my","nuclearmalaysia.gov.my","ptpk.gov.my","mohr.gov.my","driving-school.com.my","esyariah.gov.my","scpparking.com.my","giatmara.edu.my","kemas.gov.my","gundam.my","psis.edu.my","squarelet.com","xl-shop.com","kwzone.com","muftiwp.gov.my","kentrade.com.my","lagubestbest.com","polinilai.edu.my","satugadget.com.my","itradecimb.com","keda.gov.my","jawi.gov.my","water.gov.my","jmm.gov.my","flashbynight.com","ylcamera.com.my","mpsp.gov.my","mchb.com.my","unidex.com.my","idealtech.com.my","met.gov.my","sofrehkhune.com","nre.gov.my","ipmart.com.my","hotelchinatown2.com","melaka.gov.my","sumitrakanfk.com","mygrants.gov.my","seraphim.my","jpph.gov.my","tuneprotect.com","redtone.com","guvnl.com","waktusolat.net","dongzong.my","tsunjin.edu.my","ohlirik.com","seiketsu.com.my","pkb.edu.my","skyprint.com.my","kpwkm.gov.my","renren.com"]

counter = len(icp_test_pages)

wb = openpyxl.Workbook()
ws = wb.active

for i in range(counter):

    ## ADD A SECTION TO MODIFY THE DOMAIN FOR 6 SCENARIOS ##

    test_page1 = "http://" + icp_test_pages[i] 
    test_page2 = "https://" + icp_test_pages[i]
    test_page3 = "http://" + icp_test_pages[i] + ".cn"
    test_page4 = "https://" + icp_test_pages[i] + ".cn"
    test_page5 = "http://" + icp_test_pages[i].split(".")[0] + ".cn"
    test_page6 = "https://" + icp_test_pages[i].split(".")[0] + ".cn"

    test_page_list = [test_page1,test_page2,test_page3,test_page4,test_page5,test_page6]

    ########################################################

    ws.cell(row = 1, column = 1).value ="Domain" 
    ws.cell(row = 1, column = 2).value ="ICP (Y/N)" 
    #ws.cell(row = 1, column = 3).value ="Found on these domains:"
    

    #ws.cell(row = 1, column = 2).value ="Domain" 
    #len(ws['C'])

    #ws.cell(row = len(ws['A']), column = 1).value = icp_test_pages[i]

    for x in range (6):        
        test_page = test_page_list[x]
        print(test_page)

        try:
            page = requests.get(test_page)
            soup = BeautifulSoup(page.content,"html.parser")

            if (bool(soup.find_all(text=re.compile("ICP"))) == True):
                ws.cell(row = len(ws['A'])+1, column = 1).value = test_page
                ws.cell(row = len(ws['A']), column = 2).value = "Y"
                print ("Most likely has an ICP")
                wb.save("C:/Users/Florian Parzhuber/Desktop/ICP Check/Malaysia1000.xlsx")
                #print (soup.find_all(text=re.compile("京ICP证")))    
            else:
                #ws.cell(row = len(ws['A']), column = 2).value = "N"
                print("Most likely has no ICP")
    
        except:
            print("invalid domain, continue")


wb.save("C:/Users/Florian Parzhuber/Desktop/ICP Check/Malaysia100.xlsx")



